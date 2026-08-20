from __future__ import annotations

from dataclasses import FrozenInstanceError
from datetime import date
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook

from westbusan.vacant_house import source
from westbusan.vacant_house.models import VacantHouseSourceError
from westbusan.vacant_house.source import iter_archive_rows, profile_archive

REQUIRED_HEADERS = (
    "시군구코드",
    "읍면동코드",
    "시군구",
    "읍면동",
    "토지구분",
    "본번",
    "부번",
    "도로명주소",
    "건축연도",
    "무허가여부",
    "철거필요여부",
    "빈집등급",
)
PRIVATE_ADDRESS = "PRIVATE-ROAD-ADDRESS-777"
PRIVATE_WORKBOOK_NAME = "private-modern-source.xlsx"
PRIVATE_SHEET_NAME = "Private Source Sheet"


def _source_values(district_code: object = "26380") -> list[object]:
    return [
        district_code,
        "10100",
        "테스트구",
        "테스트동",
        "대지",
        12,
        3,
        PRIVATE_ADDRESS,
        1999,
        0,
        1,
        "2등급",
    ]


def _xlsx_bytes(
    rows: list[list[object]], *, sheet_name: str = PRIVATE_SHEET_NAME
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _write_archive(path: Path, members: list[tuple[str, bytes]]) -> None:
    with ZipFile(path, "w", compression=ZIP_DEFLATED) as archive:
        for name, raw in members:
            if name.endswith("/"):
                archive.writestr(name, b"")
            else:
                archive.writestr(name, raw)


class _FakeLegacySheet:
    name = "Private Legacy Sheet"

    def __init__(self) -> None:
        self._rows = [list(REQUIRED_HEADERS), _source_values()]
        self.nrows = len(self._rows)

    def row_values(self, row_index: int) -> list[object]:
        return self._rows[row_index]


class _FakeLegacyWorkbook:
    nsheets = 1

    def __init__(self) -> None:
        self.released = False

    def sheet_by_index(self, sheet_index: int) -> _FakeLegacySheet:
        assert sheet_index == 0
        return _FakeLegacySheet()

    def release_resources(self) -> None:
        self.released = True


def test_profile_counts_mixed_workbook_formats_and_ignores_directory_member(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A wrong magic classifier or directory filter changes the 15/1 counts."""
    modern = _xlsx_bytes([list(REQUIRED_HEADERS), _source_values()])
    legacy = bytes.fromhex("D0CF11E0A1B11AE1") + b"legacy-test-content"
    archive_path = tmp_path / "source.zip"
    members = [("ignored-directory.xlsx/", b"")]
    members.extend((f"nested/modern-{index}.xlsx", modern) for index in range(15))
    members.append(("nested/legacy-disguised-as-modern.xlsx", legacy))
    _write_archive(archive_path, members)

    opened: list[tuple[bytes, bool]] = []

    def fake_open_workbook(*, file_contents: bytes, on_demand: bool):
        opened.append((file_contents, on_demand))
        return _FakeLegacyWorkbook()

    monkeypatch.setattr(source.xlrd, "open_workbook", fake_open_workbook)

    profile = profile_archive(archive_path)

    assert profile.workbook_count == 16
    assert profile.modern_workbook_count == 15
    assert profile.legacy_workbook_count == 1
    assert profile.sheet_count == 16
    assert profile.candidate_row_count == 16
    assert profile.archive_sha256 == sha256(archive_path.read_bytes()).hexdigest()
    assert opened == [(legacy, True)]
    with pytest.raises(FrozenInstanceError):
        profile.workbook_count = 0  # type: ignore[misc]


def test_iterates_multiline_headers_with_hashed_labels_and_immutable_values(
    tmp_path: Path,
) -> None:
    """Whitespace/notes must not hide required headers or leak raw labels."""
    headers = [f"{header[:2]}\n{header[2:]} (작성 지침)" for header in REQUIRED_HEADERS]
    raw = _xlsx_bytes([headers, _source_values()])
    archive_path = tmp_path / "source.zip"
    member_name = f"nested/{PRIVATE_WORKBOOK_NAME}"
    _write_archive(archive_path, [("nested/", b""), (member_name, raw)])

    rows = list(iter_archive_rows(archive_path, date(2025, 2, 28)))

    assert len(rows) == 1
    row = rows[0]
    assert row.district_code == "26380"
    assert row.source_row_number == 2
    assert row.source_format == "xlsx"
    assert set(REQUIRED_HEADERS) <= set(row.values)
    assert row.workbook_sha256 == sha256(raw).hexdigest()
    assert row.workbook_name_hash == sha256(member_name.encode()).hexdigest()
    assert row.sheet_name_hash == sha256(PRIVATE_SHEET_NAME.encode()).hexdigest()
    assert PRIVATE_WORKBOOK_NAME not in repr(row)
    assert PRIVATE_SHEET_NAME not in repr(row)
    with pytest.raises(TypeError):
        row.values["시군구코드"] = "99999"  # type: ignore[index]
    with pytest.raises(FrozenInstanceError):
        row.district_code = "99999"  # type: ignore[misc]


def test_two_row_grouped_headers_preserve_excel_source_row_number(
    tmp_path: Path,
) -> None:
    """Using only one header row would reject valid grouped headers."""
    grouped = (
        ("시군구", "코드"),
        ("읍면동", "코드"),
        ("시군구", ""),
        ("읍면동", ""),
        ("토지", "구분"),
        ("본", "번"),
        ("부", "번"),
        ("도로명", "주소"),
        ("건축", "연도"),
        ("무허가", "여부"),
        ("철거필요", "여부"),
        ("빈집", "등급"),
    )
    raw = _xlsx_bytes(
        [
            ["private title"],
            [top for top, _ in grouped],
            [bottom for _, bottom in grouped],
            _source_values(26380.0),
        ]
    )
    archive_path = tmp_path / "source.zip"
    _write_archive(archive_path, [(PRIVATE_WORKBOOK_NAME, raw)])

    rows = list(iter_archive_rows(archive_path, date(2025, 2, 28)))

    assert len(rows) == 1
    assert rows[0].district_code == "26380"
    assert rows[0].source_row_number == 4
    assert set(rows[0].values) == set(REQUIRED_HEADERS)


def test_reads_legacy_workbook_through_xlrd_on_demand(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Dispatching legacy magic to the modern reader would omit its rows."""
    legacy = bytes.fromhex("D0CF11E0A1B11AE1") + b"legacy-test-content"
    archive_path = tmp_path / "source.zip"
    member_name = "nested/private-legacy-source.xlsx"
    _write_archive(archive_path, [(member_name, legacy)])
    workbook = _FakeLegacyWorkbook()
    calls: list[tuple[bytes, bool]] = []

    def fake_open_workbook(*, file_contents: bytes, on_demand: bool):
        calls.append((file_contents, on_demand))
        return workbook

    monkeypatch.setattr(source.xlrd, "open_workbook", fake_open_workbook)

    rows = list(iter_archive_rows(archive_path, date(2025, 2, 28)))

    assert len(rows) == 1
    assert rows[0].source_format == "xls"
    assert rows[0].source_row_number == 2
    assert rows[0].workbook_name_hash == sha256(member_name.encode()).hexdigest()
    assert calls == [(legacy, True)]
    assert workbook.released is True


def test_mixed_district_sheet_fails_closed_without_sensitive_error_text(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """Failing to compare all row districts would admit a cross-district sheet."""
    raw = _xlsx_bytes(
        [list(REQUIRED_HEADERS), _source_values("26380"), _source_values("26440")]
    )
    archive_path = tmp_path / "source.zip"
    _write_archive(archive_path, [(PRIVATE_WORKBOOK_NAME, raw)])

    with pytest.raises(VacantHouseSourceError) as caught:
        list(iter_archive_rows(archive_path, date(2025, 2, 28)))

    output = capsys.readouterr()
    assert caught.value.code == "mixed_district_sheet"
    assert str(caught.value) == "mixed_district_sheet"
    assert PRIVATE_ADDRESS not in str(caught.value)
    assert PRIVATE_WORKBOOK_NAME not in str(caught.value)
    assert PRIVATE_SHEET_NAME not in str(caught.value)
    assert output.out == ""
    assert output.err == ""


def test_unknown_workbook_magic_is_rejected_with_safe_code(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """Extension-only detection would pass unsupported workbook bytes downstream."""
    archive_path = tmp_path / "source.zip"
    _write_archive(
        archive_path,
        [(PRIVATE_WORKBOOK_NAME, b"unsupported private workbook content")],
    )

    with pytest.raises(VacantHouseSourceError) as caught:
        profile_archive(archive_path)

    output = capsys.readouterr()
    assert caught.value.code == "unsupported_workbook_format"
    assert str(caught.value) == "unsupported_workbook_format"
    assert PRIVATE_WORKBOOK_NAME not in str(caught.value)
    assert output.out == ""
    assert output.err == ""


def test_missing_required_headers_are_rejected_without_source_labels(
    tmp_path: Path,
) -> None:
    """Accepting a partial schema would silently turn missing evidence into blanks."""
    raw = _xlsx_bytes([["시군구코드", "도로명주소"], ["26380", PRIVATE_ADDRESS]])
    archive_path = tmp_path / "source.zip"
    _write_archive(archive_path, [(PRIVATE_WORKBOOK_NAME, raw)])

    with pytest.raises(VacantHouseSourceError) as caught:
        list(iter_archive_rows(archive_path, date(2025, 2, 28)))

    assert caught.value.code == "required_headers_missing"
    assert str(caught.value) == "required_headers_missing"
    assert PRIVATE_ADDRESS not in str(caught.value)
    assert PRIVATE_WORKBOOK_NAME not in str(caught.value)
    assert PRIVATE_SHEET_NAME not in str(caught.value)
