"""Immutable ingestion of manually supplied transport evidence files."""

from __future__ import annotations

import hashlib
import re
from dataclasses import replace
from datetime import date
from pathlib import Path
from uuid import NAMESPACE_URL, uuid5

from openpyxl import load_workbook

from westbusan.models import RawArtifact, RunContext
from westbusan.storage import RawStore

_SUPPORTED = {".csv", ".xlsx"}
_FILENAME_RULES = {
    "korail_workplace_ticketing_file": (("korail", "한국철도공사"), ("근무",)),
    "korail_residence_ticketing_file": (("korail", "한국철도공사"), ("거주",)),
    "srt_station_boarding_file": (("srt", "에스알"), ("역", "승하차", "승차", "하차")),
}


class FileSource:
    """Stores original CSV/XLSX evidence before any tabular interpretation."""

    def __init__(self, data_dir: Path) -> None:
        self.store = RawStore(data_dir)

    def ingest(self, path: Path, source_id: str, run: RunContext) -> RawArtifact:
        """Copy one supported evidence file into immutable content-addressed storage."""
        path = Path(path)
        if path.suffix.lower() not in _SUPPORTED:
            raise ValueError("transport files must be CSV or XLSX")
        if not path.is_file():
            raise FileNotFoundError(path)
        body = path.read_bytes()
        source_date, source_date_quality, source_date_granularity, source_date_value = _source_date(path)
        artifact = self.store.write(
            run,
            source_id,
            {
                "kind": "file",
                "filename": path.name,
                "content_hash": file_fingerprint(path),
                "source_date_quality": source_date_quality,
                "source_date_granularity": source_date_granularity,
                "source_date_value": source_date_value,
            },
            body,
            path.suffix.lower(),
            source_date=source_date,
        )
        # A content-addressed byte path is deliberately shared, but every pipeline
        # run receives an artifact row so a repeated file remains auditable.
        return replace(
            artifact,
            artifact_id=uuid5(
                NAMESPACE_URL,
                f"file:{source_id}:{run.run_id}:{artifact.request_hash}:{artifact.content_hash}",
            ),
        )

    def discover(self, inbox: Path, source_id: str) -> tuple[Path, ...]:
        """Return only files whose names match the approved source-specific pattern."""
        rule = _FILENAME_RULES.get(source_id)
        if rule is None:
            raise KeyError(f"no approved filename pattern for {source_id}")
        if not inbox.exists():
            return ()
        return tuple(
            path
            for path in sorted(inbox.iterdir())
            if path.is_file()
            and path.suffix.lower() in _SUPPORTED
            and _matches_filename(path.name, rule)
        )


def file_fingerprint(path: Path) -> str:
    """Return the SHA-256 content identity of a provided evidence file."""
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()


def read_tabular_rows(path: Path) -> list[dict[str, object]]:
    """Read a supported evidence file without shifting blanks or string cells."""
    path = Path(path)
    if path.suffix.lower() == ".csv":
        return _csv_rows(path)
    if path.suffix.lower() == ".xlsx":
        return _xlsx_rows(path)
    raise ValueError("transport files must be CSV or XLSX")


def _csv_rows(path: Path) -> list[dict[str, object]]:
    import csv

    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "cp949"):
        try:
            with path.open(encoding=encoding, newline="") as stream:
                reader = csv.DictReader(stream)
                headers = _headers(reader.fieldnames)
                return [{name: row.get(name) for name in headers} for row in reader]
        except UnicodeDecodeError as error:
            last_error = error
    raise ValueError(f"CSV is neither UTF-8 nor CP949: {path.name}") from last_error


def _xlsx_rows(path: Path) -> list[dict[str, object]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        headers = _headers(next(iterator, None))
        rows = [dict(zip(headers, values, strict=True)) for values in iterator]
        workbook.close()
        return rows
    except Exception as error:
        raise ValueError(f"unable to read XLSX: {path.name}") from error


def _headers(values: object) -> tuple[str, ...]:
    if not isinstance(values, (list, tuple)) or not values:
        raise ValueError("tabular file has no header row")
    headers = tuple(str(value).strip() if value is not None else "" for value in values)
    if not all(headers) or len(headers) != len(set(headers)):
        raise ValueError("tabular file headers must be non-empty and unique")
    return headers


def _matches_filename(name: str, rule: tuple[tuple[str, ...], tuple[str, ...]]) -> bool:
    provider_terms, subject_terms = rule
    normalized = name.lower()
    return any(term in normalized for term in provider_terms) and any(
        term in normalized for term in subject_terms
    )


def _source_date(path: Path) -> tuple[date | None, str, str, str | None]:
    match = re.search(r"(?<!\d)((?:19|20)\d{2})(?:[-_.]?(\d{2}))?(?:[-_.]?(\d{2}))?", path.name)
    if match:
        year, month, day = match.groups()
        if month is None:
            return None, "filename", "year", year
        try:
            return (
                date(int(year), int(month), int(day or 1)),
                "filename",
                "day" if day is not None else "month",
                f"{year}-{month}-{day}" if day is not None else f"{year}-{month}",
            )
        except ValueError:
            pass
    return None, "unknown", "unknown", None


__all__ = ["FileSource", "file_fingerprint", "read_tabular_rows"]
