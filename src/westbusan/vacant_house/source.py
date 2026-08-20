"""Read-only parsing for mixed-format vacant-house source archives."""

from __future__ import annotations

import re
from collections.abc import Iterator, Mapping, Sequence
from dataclasses import dataclass
from datetime import date
from hashlib import sha256
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Literal
from zipfile import BadZipFile, ZipFile, ZipInfo

import xlrd
from openpyxl import load_workbook

from westbusan.vacant_house.models import (
    ArchiveProfile,
    VacantHouseSourceError,
    VacantHouseSourceRow,
)

XLSX_MAGIC = b"PK\x03\x04"
XLS_MAGIC = bytes.fromhex("D0CF11E0A1B11AE1")
REQUIRED_HEADERS = frozenset(
    {
        "시군구코드",
        "읍면동코드",
        "시군구",
        "읍면동",
        "토지구분",
        "본번",
        "부번",
        "도로명주소",
        "건축연도",
        "무허가여부",
        "철거필요여부",
        "빈집등급",
    }
)
_WORKBOOK_SUFFIXES = frozenset({".xlsx", ".xls"})
_HEADER_SCAN_LIMIT = 50
_PARENTHETICAL = re.compile(r"\([^()]*\)|（[^（）]*）")
_HEADER_ALIASES = {
    "번지구분": "토지구분",
    "본번*숫자만": "본번",
    "부번*숫자만": "부번",
    "빈집": "빈집등급",
    "건축년도": "건축연도",
}


@dataclass(frozen=True, slots=True)
class _WorkbookEntry:
    raw: bytes
    name_hash: str
    source_format: Literal["xlsx", "xls"]


@dataclass(frozen=True, slots=True)
class _ParsedSheet:
    sheet_name_hash: str
    rows: tuple[tuple[int, Mapping[str, object], str], ...]


def profile_archive(path: Path) -> ArchiveProfile:
    """Return aggregate format and candidate-row evidence for an archive."""
    archive_sha256, entries = _read_archive(Path(path))
    modern_workbook_count = 0
    legacy_workbook_count = 0
    sheet_count = 0
    candidate_row_count = 0

    for entry in entries:
        if entry.source_format == "xlsx":
            modern_workbook_count += 1
        else:
            legacy_workbook_count += 1
        parsed_sheets = _parse_workbook(entry)
        sheet_count += len(parsed_sheets)
        candidate_row_count += sum(len(sheet.rows) for sheet in parsed_sheets)

    return ArchiveProfile(
        archive_sha256=archive_sha256,
        workbook_count=len(entries),
        modern_workbook_count=modern_workbook_count,
        legacy_workbook_count=legacy_workbook_count,
        sheet_count=sheet_count,
        candidate_row_count=candidate_row_count,
    )


def iter_archive_rows(
    path: Path, snapshot_date: date
) -> Iterator[VacantHouseSourceRow]:
    """Yield immutable private rows without changing the source archive."""
    if not isinstance(snapshot_date, date):
        raise TypeError("snapshot_date must be a date")

    _, entries = _read_archive(Path(path))
    for entry in entries:
        workbook_sha256 = sha256(entry.raw).hexdigest()
        for sheet in _parse_workbook(entry):
            for source_row_number, values, district_code in sheet.rows:
                yield VacantHouseSourceRow(
                    workbook_sha256=workbook_sha256,
                    workbook_name_hash=entry.name_hash,
                    sheet_name_hash=sheet.sheet_name_hash,
                    source_row_number=source_row_number,
                    source_format=entry.source_format,
                    values=values,
                    district_code=district_code,
                )


def _read_archive(path: Path) -> tuple[str, tuple[_WorkbookEntry, ...]]:
    try:
        raw_archive = path.read_bytes()
        with ZipFile(BytesIO(raw_archive)) as archive:
            entries = tuple(
                _read_workbook_entry(archive, info)
                for info in archive.infolist()
                if _is_workbook_member(info)
            )
    except VacantHouseSourceError:
        raise
    except (BadZipFile, OSError, RuntimeError, ValueError):
        raise VacantHouseSourceError("invalid_archive") from None
    return sha256(raw_archive).hexdigest(), entries


def _is_workbook_member(info: ZipInfo) -> bool:
    if info.is_dir():
        return False
    return PurePosixPath(info.filename).suffix.lower() in _WORKBOOK_SUFFIXES


def _read_workbook_entry(archive: ZipFile, info: ZipInfo) -> _WorkbookEntry:
    raw = archive.read(info)
    return _WorkbookEntry(
        raw=raw,
        name_hash=sha256(info.filename.encode("utf-8")).hexdigest(),
        source_format=_detect_workbook_format(raw),
    )


def _detect_workbook_format(raw: bytes) -> Literal["xlsx", "xls"]:
    if raw.startswith(XLSX_MAGIC):
        return "xlsx"
    if raw.startswith(XLS_MAGIC):
        return "xls"
    raise VacantHouseSourceError("unsupported_workbook_format")


def _parse_workbook(entry: _WorkbookEntry) -> tuple[_ParsedSheet, ...]:
    raw_sheets = (
        _read_xlsx_sheets(entry.raw)
        if entry.source_format == "xlsx"
        else _read_xls_sheets(entry.raw)
    )
    return tuple(_parse_sheet(sheet_name, rows) for sheet_name, rows in raw_sheets)


def _read_xlsx_sheets(
    raw: bytes,
) -> tuple[tuple[str, tuple[tuple[object, ...], ...]], ...]:
    workbook = None
    try:
        workbook = load_workbook(
            filename=BytesIO(raw),
            read_only=True,
            data_only=True,
        )
        return tuple(
            (
                sheet.title,
                tuple(tuple(row) for row in sheet.iter_rows(values_only=True)),
            )
            for sheet in workbook.worksheets
        )
    except Exception:  # noqa: BLE001 - replace untrusted parser details with safe code
        raise VacantHouseSourceError("unreadable_workbook") from None
    finally:
        if workbook is not None:
            workbook.close()


def _read_xls_sheets(
    raw: bytes,
) -> tuple[tuple[str, tuple[tuple[object, ...], ...]], ...]:
    workbook = None
    try:
        workbook = xlrd.open_workbook(file_contents=raw, on_demand=True)
        return tuple(
            (
                sheet.name,
                tuple(tuple(sheet.row_values(row_index)) for row_index in range(sheet.nrows)),
            )
            for sheet in (
                workbook.sheet_by_index(index) for index in range(workbook.nsheets)
            )
        )
    except Exception:  # noqa: BLE001 - replace untrusted parser details with safe code
        raise VacantHouseSourceError("unreadable_workbook") from None
    finally:
        if workbook is not None:
            workbook.release_resources()


def _parse_sheet(
    sheet_name: str, rows: Sequence[Sequence[object]]
) -> _ParsedSheet:
    header_end, headers = _locate_headers(rows)
    parsed_rows: list[tuple[int, Mapping[str, object], str]] = []
    district_codes: set[str] = set()
    district_names: set[str] = set()

    for source_row_number, row in enumerate(
        rows[header_end + 1 :], start=header_end + 2
    ):
        if _is_empty_row(row):
            continue
        values = {
            header: row[column_index] if column_index < len(row) else None
            for column_index, header in headers.items()
        }
        district_code = _source_code(values["시군구코드"])
        if not district_code:
            continue
        district_name = _source_text(values["시군구"])
        district_codes.add(district_code)
        if district_name:
            district_names.add(district_name)
        parsed_rows.append((source_row_number, values, district_code))

    if len(district_codes) > 1 or len(district_names) > 1:
        raise VacantHouseSourceError("mixed_district_sheet")

    return _ParsedSheet(
        sheet_name_hash=sha256(sheet_name.encode("utf-8")).hexdigest(),
        rows=tuple(parsed_rows),
    )


def _locate_headers(
    rows: Sequence[Sequence[object]],
) -> tuple[int, Mapping[int, str]]:
    last_start = min(len(rows), _HEADER_SCAN_LIMIT)
    for start in range(last_start):
        for height in (1, 2):
            end = start + height
            if end > len(rows):
                continue
            headers = _map_headers(rows[start:end])
            if REQUIRED_HEADERS <= set(headers.values()):
                return end - 1, headers
    raise VacantHouseSourceError("required_headers_missing")


def _map_headers(header_rows: Sequence[Sequence[object]]) -> dict[int, str]:
    column_count = max((len(row) for row in header_rows), default=0)
    headers: dict[int, str] = {}
    for column_index in range(column_count):
        parts = [
            _normalize_header(row[column_index])
            for row in header_rows
            if column_index < len(row)
        ]
        parts = [part for part in parts if part]
        if not parts:
            continue
        combined = "".join(dict.fromkeys(parts))
        header = combined if combined in REQUIRED_HEADERS else _required_part(parts)
        headers[column_index] = header or combined
    return headers


def _required_part(parts: Sequence[str]) -> str | None:
    return next((part for part in reversed(parts) if part in REQUIRED_HEADERS), None)


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    previous = None
    while previous != text:
        previous = text
        text = _PARENTHETICAL.sub("", text)
    normalized = "".join(text.split())
    return _HEADER_ALIASES.get(normalized, normalized)


def _is_empty_row(row: Sequence[object]) -> bool:
    return all(value is None or (isinstance(value, str) and not value.strip()) for value in row)


def _source_code(value: object) -> str:
    if isinstance(value, bool) or value is None:
        return "" if value is None else str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _source_text(value: object) -> str:
    return "" if value is None else "".join(str(value).split())
